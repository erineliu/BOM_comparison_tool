import configparser
from openpyxl import load_workbook,Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import pandas as pd
import sys,os
import re

global c

def get_crb_section_data(section_name):
    section_data_list=[]
    for row in c.options(section_name):
        if row.find("\t") != -1:
            section_data_list.append(row.split("\t"))
        else:
            section_data_list.append(row.split())

       
    return section_data_list


def crb_to_dataframe(crb_flie):
    global c

    c = configparser.ConfigParser(allow_no_value=True)
    c.read(crb_flie)

    data_list = get_crb_section_data("PartsData")
    ## data frame ##
    #print(data_list[0])
    df = pd.DataFrame(data_list[1:],columns=data_list[0])
    part2_df = df[['idnum','name','fa', 'pack']]
    #print(part2_df)

    

    data_list = get_crb_section_data("PositionData<1>")
    df = pd.DataFrame(data_list[1:],columns=data_list[0])
    part1_df = df[['idnum','parts','c']]
    # print(part2_df)


    data_list = get_crb_section_data("ChipData")
    df = pd.DataFrame(data_list[1:],columns=data_list[0])
    part3_df = df[['idnum','l','w','t']]


    data_list = get_crb_section_data("ShapeBase")
    df = pd.DataFrame(data_list[1:],columns=data_list[0])
    part5_df = df[['idnum','chipl','chipw','chiph']]
    #print(part5_df)

    #sys.exit(1)

    data_list = get_crb_section_data("FeederData")
    df = pd.DataFrame(data_list[1:],columns=data_list[0])
    part4_df = df[['idnum','feedlen','width']]
    #print(part4_df)

    
    part1_df = part1_df.reindex(columns = part1_df.columns.tolist() + ['name','fa','pack',"l","w","t","feedlen","width"])
    

    for index, row in part1_df.iterrows():

       
        
        ## find in part2_df ##
        match2_df = part2_df.loc[part2_df['idnum'] == row['parts']]
        if not match2_df.empty:
          
            part1_df.loc[ index, ["name","fa","pack"]] = [ match2_df["name"].values[0], match2_df["fa"].values[0], match2_df["pack"].values[0] ]
            #print(part1_df)
            #sys.exit(1)

            ## find in part3_df ##
            match3_df = part3_df.loc[part3_df['idnum'] == match2_df['idnum'].values[0]]
            if not match3_df.empty:
                part1_df.loc[index, ['l','w','t'] ] = [ match3_df["l"].values[0], match3_df["w"].values[0], match3_df["t"].values[0] ]


            ## find in part5_df ##
            match5_df = part5_df.loc[part5_df['idnum'] == match2_df['idnum'].values[0]]
            if not match5_df.empty:
                #print(match5_df["chipl"].values[0], match5_df["chipw"].values[0], match5_df["chiph"].values[0])
                part1_df.loc[index, ['l','w','t'] ] = [ match5_df["chipl"].values[0], match5_df["chipw"].values[0], match5_df["chiph"].values[0] ]

                #print(part1_df)
                #sys.exit(1)

        ## find in part4_df ##
        #print(row['idnum'])
        #print(part4_df)
        
            match4_df = part4_df.loc[part4_df['idnum'] == match2_df["fa"].values[0]]
           
            if not match4_df.empty:
                part1_df.loc[index, ['feedlen','width'] ] = [ match4_df["feedlen"].values[0], match4_df["width"].values[0] ]

    
    #print(part1_df)
    return part1_df

## ============================================================= ##

## output format ## 

#header1 = ["location","Parts NO","fa","parts","pack","l","w","t","feedlen","width"]
#header2 = ["Parts NO2","fa2","parts2","pack2","l2","w2","t2","feedlen2","width2"]
header1 = ["Location","Parts NO","pack","L","W","T","Feedlen","Width"]
header2 = ["Parts NO2","Pack","L","W","T","Feedlen","Width"]
result = ["Result"]

header = header1 + header2 + result
output_df = pd.DataFrame(columns = header)



## ============================================================= ##
r,d,f = next(os.walk("input"))
r = re.compile("MAIN*")
main_list = list(filter(r.match,f))


main_df = pd.DataFrame()
for index,cur_main_file in enumerate(main_list):
    cur_main_df = crb_to_dataframe("input/" + cur_main_file)
    if index == 0:
        main_df = cur_main_df
    else:
        main_df = main_df.append(cur_main_df, ignore_index=True)



main_df = main_df[ main_df.c != '\"\"' ]

#main_B_df = crb_to_dataframe("02201BAA-17B08.crb")
#main_T_df = crb_to_dataframe("02201TAA-17T08.crb")
#main_df = main_B_df.append(main_T_df, ignore_index=True)

at_list = []
r = re.compile("^(?!MAIN).*")
at_input_list = list(filter(r.match,f))
for index,cur_at_file in enumerate(at_input_list):
    at_df = crb_to_dataframe("input/" + cur_at_file )
    at_list.append(at_df)

print(at_input_list)
#print(at_list)



#at1_df = crb_to_dataframe("input/" + "02201TAA-17T10.crb")
#at2_df = crb_to_dataframe("input/" + "02201BAA-17B09.crb")
#print(main_B_df)
#print(main_T_df)
#print(main_df)
#print(at1_df)
#print(at2_df)
#at_list = [at1_df,at2_df]


# ## =========================== ##
# writer = pd.ExcelWriter('main_input.xlsx')
# ## Convert the dataframe to an XlsxWriter Excel object. ##
# main_df.to_excel(writer, sheet_name='Sheet1')
# ## Close the Pandas Excel writer and output the Excel file. ##
# writer.save()
# #sys.exit(1)

# ## =========================== ##

for index, row in main_df.iterrows():
    #print(main_df.loc[0,["c","l"]].values)
    
    main_row = row[["fa","pack","l","w","t","feedlen","width"]].values
    main_row2 = [ float(i) for i in main_row ]
    #print(main_row2)
    for index,cur_at_df in enumerate(at_list):

        #match_at_df = cur_at_df.loc[cur_at_df['c'] == '\"%s\"'%(row['c'])]
        match_at_df = cur_at_df.loc[cur_at_df['c'] == row['c']]


        if not match_at_df.empty:
            #print(match_at_df,index)

            alt_row = match_at_df[["fa","pack","l","w","t","feedlen","width"]].values[0]
            alt_row2 = [float(i) for i in alt_row]

            if main_row2 == alt_row2:
                result = "PASS"
            else:
                #print(main_row2,alt_row2)
                result = "FAIL" 
        
            #if True:
            if result == "FAIL": 
                output_loc_row = list(row[["c"]].values)
                #output_main_row = list(row[["name","fa","parts","pack","l","w","t","feedlen","width"]].values)
                #output_alt_row =list(match_at_df[["name","fa","parts","pack","l","w","t","feedlen","width"]].values[0])
                output_main_row = list(row[["name","pack","l","w","t","feedlen","width"]].values)
                output_alt_row =list(match_at_df[["name","pack","l","w","t","feedlen","width"]].values[0])

                output_data = output_loc_row + output_main_row + output_alt_row + [result] 
                output_df.loc[len(output_df.index)] = output_data
                #output_df = output_df.append(pd.Series(output_data, index=output_df.columns ), ignore_index=True)



writer = pd.ExcelWriter('output_PSE.xlsx')
## Convert the dataframe to an XlsxWriter Excel object. ##
output_df.to_excel(writer, sheet_name='Sheet1')
## Close the Pandas Excel writer and output the Excel file. ##
writer.save()


## ============================================================= ##

wb3 = load_workbook('output_PSE.xlsx')

first_sheet = wb3.get_sheet_names()[0]
cur_sheet = wb3.get_sheet_by_name(first_sheet)

head_font = Font(name='Microsoft YaHei',bold=True)
data_font = Font(name='Microsoft YaHei')
for cell in cur_sheet[1:1]:
    cell.font = head_font

#print(cur_sheet.max_row)
#print(cur_sheet['A2':'B8'])

for i in range(2,cur_sheet.max_row+1):
    for cell in cur_sheet[i:i]:
        cell.font = data_font

## ajuest the column width ##
for col in cur_sheet.columns:
    max_length = 0
    column = col[0].column_letter   # Get the column name A,B,C,D..
    #if column == "A":
    #    continue
   
    for cell in col:
        
        if cell.coordinate in cur_sheet.merged_cells: # not check merge_cells
            continue
       
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
        
    adjusted_width = (max_length + 2) * 1.2
    cur_sheet.column_dimensions[column].width = adjusted_width
    wb3.save("output_PSE.xlsx")